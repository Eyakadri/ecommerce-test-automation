import json
import os
from openpyxl import load_workbook
from faker import Faker


class DataLoader:
    """Handles loading test data from JSON and Excel files."""
    
    @staticmethod
    def load_json(filepath):
        """
        Load data from a JSON file.
        
        Args:
            filepath: Path to the JSON file
            
        Returns:
            Parsed JSON data (dict or list)
        """
        try:
            with open(filepath, 'r', encoding='utf-8') as file:
                return json.load(file)
        except FileNotFoundError:
            print(f"✗ JSON file not found: {filepath}")
            return None
        except json.JSONDecodeError as e:
            print(f"✗ Error parsing JSON: {str(e)}")
            return None
    
    @staticmethod
    def load_excel(filepath, sheet_name=0):
        """
        Load data from an Excel file.
        
        Args:
            filepath: Path to the Excel file
            sheet_name: Sheet index or name (default: first sheet)
            
        Returns:
            List of dictionaries (each row mapped to column headers)
        """
        try:
            workbook = load_workbook(filepath)
            worksheet = workbook[sheet_name] if isinstance(sheet_name, str) else workbook.sheetnames[sheet_name]
            
            headers = []
            data = []
            
            for i, row in enumerate(worksheet.iter_rows(values_only=True)):
                if i == 0:
                    headers = row
                else:
                    if any(cell is not None for cell in row):  # Skip empty rows
                        data.append(dict(zip(headers, row)))
            
            return data
        except FileNotFoundError:
            print(f"✗ Excel file not found: {filepath}")
            return None
        except Exception as e:
            print(f"✗ Error reading Excel: {str(e)}")
            return None
    
    @staticmethod
    def generate_user_data(count=1):
        """
        Generate random user data using Faker.
        
        Args:
            count: Number of users to generate
            
        Returns:
            List of dictionaries with user data
        """
        fake = Faker()
        users = []
        
        for _ in range(count):
            users.append({
                'first_name': fake.first_name(),
                'last_name': fake.last_name(),
                'email': fake.email(),
                'password': fake.password(length=12, special_chars=True),
                'phone': fake.phone_number(),
                'company': fake.company(),
                'address': fake.address(),
                'city': fake.city(),
                'postcode': fake.postcode(),
                'country': fake.country()
            })
        
        return users if count > 1 else users[0]
